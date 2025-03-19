"""
该模块用于从指定的URL获取部落和玩家信息,并将其更新到Excel文件中。模块主要功能包括:
1. updateInformation: 更新部落和玩家信息。
2. deleteAll: 删除输出文件。
3. queryContribution: 查询部落战贡献信息。
4. queryDonation: 查询部落捐赠信息。
5. queryActivity: 查询玩家活动信息。
6. queryLastMonthWar: 查询上个月的部落战信息。
7. queryLastMonthDonation: 查询上个月的捐赠信息。
8. queryAndSort: 查询并排序贡献和捐赠信息。
9. queryRecentChange: 查询最近的成员变动信息。
模块依赖于以下库和模块:
- urllib.request: 用于发送HTTP请求。
- selenium: 用于自动化浏览器操作。
- BeautifulSoup: 用于解析HTML内容。
- openpyxl: 用于操作Excel文件。
- src.urls: 包含URL和请求头信息。
- src.operations: 包含Excel操作函数。
- src.clansInformation: 包含部落和玩家信息。
- src.formal: 包含格式化和排序函数。
- src.externs: 包含外部配置和路径信息。
模块通过调用上述函数,实现对部落和玩家信息的自动化获取、更新和处理,并将结果保存到指定的Excel文件中。

"""
import urllib.request
import src.urls as urls
from selenium import webdriver
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.edge.options import Options
from webdriver_manager.microsoft import EdgeChromiumDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import urllib
import src.operations as operations
from bs4 import BeautifulSoup
from src.clansInformation import clans
from src.clansInformation import players
import openpyxl as op
from openpyxl.styles import PatternFill
import src.formal as formal
import src.externs as externs
import os
from lxml import etree
import src.log as log
def updateInformation():
    flag = True
    log.log("INFO", "FETCH", "准备更新部落信息和玩家信息...", externs.log_path)
    log.log("INFO", "FETCH", "准备更新部落信息和玩家信息...", )
    log.log("INFO", "FETCH", "检查文件...", externs.log_path)
    if os.path.exists(externs.inputClansInformationLocation):
        log.log("INFO", "FETCH", f"文件'{externs.inputClansInformationLocation}'存在,准备读取...", externs.log_path)
        wb = op.load_workbook(externs.inputClansInformationLocation)
        ws = wb[externs.clansInformationSheetName]
        max_row = ws.max_row
        log.log("INFO", "FETCH", "文件读取完成,准备更新部落信息...", externs.log_path)
        for row in range(2,max_row + 1):
            clan = ws.cell(row = row,column = 2).value
            if clan is None:
                break#防止读过头
            clan = clan[1:]
            url_clan = urls.url_clan + clan
            requests = urllib.request.Request(url=url_clan, headers=urls.HEADERS)
            response = urllib.request.urlopen(requests)
            content = response.read().decode("utf-8")
            tree = etree.HTML(content)
            clan_name_element = tree.xpath(externs.XPaths["ClanName"])[0]
            clan_name = clan_name_element.text.strip().replace(' ', '')
            ws.cell(row = row,column = 1).value = clan_name
            log.log("INFO", "FETCH", f"< {clan_name} >更新完成！", externs.log_path)
        wb.save(filename = externs.inputClansInformationLocation)
        log.log("INFO", "FETCH", "全部部落信息更新完成！准备格式化...", externs.log_path)
        formal.processExcel(externs.inputClansInformationLocation)
        log.log("INFO", "FETCH", "格式化完成！", externs.log_path)
        log.log("INFO", "FETCH", "部落信息更新完成！", )
        log.log("INFO", "FETCH", "部落信息更新完成！", externs.log_path)
    else:
        log.log("INFO", "FETCH", f"文件'{externs.inputClansInformationLocation}'不存在", )
        log.log("INFO", "FETCH", f"文件'{externs.inputClansInformationLocation}'不存在", externs.log_path)
    if os.path.exists(externs.inputGroupPlayerInformationLocation):
        log.log("INFO", "FETCH", f"文件'{externs.inputGroupPlayerInformationLocation}'存在,准备读取...", externs.log_path)
        wb = op.load_workbook(externs.inputGroupPlayerInformationLocation)
        ws = wb[externs.groupPlayerInformationSheetName]
        max_row = ws.max_row
        log.log("INFO", "FETCH", "文件读取完成,准备玩家更新...", externs.log_path)
        for row in range(2,max_row + 1):
            player = ws.cell(row = row,column = 4).value
            if player is None:
                break#防止读过头
            player = player[1:]
            url_player = urls.url_player + player
            requests = urllib.request.Request(url = url_player,headers = urls.HEADERS)
            response = urllib.request.urlopen(requests)
            content = response.read().decode("utf-8")
            tree = etree.HTML(content)
            player_name_element = tree.xpath(externs.XPaths["PlayerName"])[0]
            player_name = player_name_element.text.strip().replace('\u200c','').replace('\u2006','')
            a = tree.xpath(externs.XPaths["PlayerClan"])
            if  a is None or len(a) == 0:
                clan = "无"
            else:
                clan = a[0].text.strip().replace(' ','')
            ws.cell(row = row,column = 1).value = clan
            if clan not in externs.clans:
                fill = PatternFill(fill_type = 'solid',fgColor = 'FF0000')
                ws.cell(row = row,column = 1).fill = fill
            else :
                fill = PatternFill(fill_type=None)
                ws.cell(row=row, column=1).fill = fill
            ws.cell(row = row,column = 2).value = player_name
            log.log("INFO", "FETCH", f"< {player_name} >更新完成！", externs.log_path)
        wb.save(filename = externs.inputGroupPlayerInformationLocation)
        log.log("INFO", "FETCH", "全部玩家信息更新完成！准备格式化...", externs.log_path)
        formal.processExcel(externs.inputGroupPlayerInformationLocation)
        log.log("INFO", "FETCH", "格式化完成！", externs.log_path)
        log.log("INFO", "FETCH", "玩家信息更新完成！", )
        log.log("INFO", "FETCH", "玩家信息更新完成！", externs.log_path)
    else:
        log.log("INFO", "FETCH", f"文件'{externs.inputGroupPlayerInformationLocation}'不存在", )
        log.log("INFO", "FETCH", f"文件'{externs.inputGroupPlayerInformationLocation}'不存在", externs.log_path)
    return flag

def deleteAll():
    flag = True
    log.log("INFO", "FETCH", "准备删除输出文件...", externs.log_path)
    if os.path.exists(externs.outputFileLocation):
        os.remove(externs.outputFileLocation)
        log.log("INFO", "FETCH", f"文件 '{externs.outputFileLocation}' 已被删除", externs.log_path)
    else:
        log.log("INFO", "FETCH", f"文件 '{externs.outputFileLocation}' 不存在", externs.log_path)
    log.log("INFO", "FETCH", "输出文件删除完成！", )
    log.log("INFO", "FETCH", "输出文件删除完成！", externs.log_path)
    return flag

def queryContribution(filter):
    flag = True
    sheet_name = externs.contributionsSheetName
    url_0 = urls.url_clan
    formal.creatSheet(operations.creat_contribution_sheet())
    log.log("INFO", "FETCH", "输出创建完成,准备启动程序...", externs.log_path)
    wb = op.load_workbook(externs.outputFileLocation)
    ws = wb[sheet_name]
    now_row = 1
    start_row = 2
    log.log("INFO", "FETCH", "程序启动,正在运行...", externs.log_path)
    for clan in clans:
        url_war_race = url_0 + clans[clan] + "/war/race"
        requests = urllib.request.Request(url = url_war_race,headers = urls.HEADERS)
        response = urllib.request.urlopen(requests)
        content = response.read().decode("utf-8")
        tree = etree.HTML(content)
        in_war = 0##可能需要别的状态,不用bool
        data_num = 0
        for day in range(1, 8):  # 第四天开始战斗日
            dot_text = tree.xpath(externs.XPaths["WarTimeline"] + "/li[" + str(day) + "]/div[3]")[0].text.strip()
            if dot_text is None or len(dot_text) == 0:
                continue
            else:
                in_war = 1
                break
        data = [clan]
        if 1 == in_war:
            log.log("INFO", "FETCH", f"< {clan} >正处于战斗日,开始查询...", externs.log_path)
            log.log("INFO", "FETCH", f"< {clan} >正处于战斗日,开始查询...")
            playersTable = tree.xpath(externs.XPaths["WarPlayersTable"])[0]
            for player in playersTable:
                player =etree.tostring(player,encoding='utf-8').decode('utf-8')
                if "player not_current_member" in player:
                    break
                player = BeautifulSoup(player,'html.parser')
                data = [clan]
                player_name =player.find('a',{'class':'player_name force_single_line_hidden'}).get_text().strip().replace('\u200c','').replace('\u2006','')
                if filter and player_name not in players:
                    continue
                data.append(player_name)
                decks_uesd = (int)(player.find('div',{'class':'value_bg decks_used'}).get_text())
                data.append(decks_uesd) 
                boat_attack = (int)(player.find('div',{'class':'value_bg boat_attacks'}).get_text())
                data.append(boat_attack)
                medal = (int)( player.find('div',{'class':'value_bg fame'}).get_text())
                data.append(medal)
                data_num = data_num + 1
                ws.append(data)
            now_row = now_row + data_num
            wb.save(filename = externs.outputFileLocation)
            formal.sort_xlsx_data(externs.outputFileLocation,sheet_name,start_row = start_row,end_row = now_row,sort_column = 5) 
            wb = op.load_workbook(externs.outputFileLocation)
            ws = wb[sheet_name]
            ws.merge_cells(start_row = start_row,end_row = now_row,start_column = 1,end_column = 1)
            start_row = now_row + 1
        elif 0 == in_war:
            log.log("INFO", "FETCH", f"< {clan} >未处于战斗日,格式输出正在进行...", externs.log_path)
            data = [clan]
            data.append('该部落未处于战斗日！')
            data.append('')
            data.append('')
            data.append('')
            ws.append(data)
            now_row = now_row + 1
            start_row = now_row
            ws.merge_cells(start_column = 2,end_column = 5,start_row = now_row,end_row = now_row)
        log.log("INFO", "FETCH", f"< {clan} >查询已完成！", externs.log_path)
    for row in ws.iter_rows(min_row = 2,max_row = ws.max_row,min_col=5,max_col=5):
        if row[0].value == 0:
            fill = PatternFill(fill_type = 'solid',fgColor = 'FF0000')
            ws.cell(row = row[0].row,column = 5).fill = fill
    wb.save(filename = externs.outputFileLocation)
    log.log("INFO", "FETCH", "部落战贡献查询完成！", )
    log.log("INFO", "FETCH", "部落战贡献查询完成！", externs.log_path)
    return flag

def queryDonation(filter):
    flag = True
    sheet_name = formal.donationSheetName
    url_0 = urls.url_clan
    formal.creatSheet(operations.creat_donation_sheet())
    log.log("INFO", "FETCH", "输出创建完成,准备启动程序...", externs.log_path)
    wb = op.load_workbook(externs.outputFileLocation)
    ws = wb[sheet_name]
    data_num = 0
    now_row = 1
    start_row = 2
    log.log("INFO", "FETCH", "程序启动,正在运行...", externs.log_path)
    for clan in clans:
        log.log("INFO", "FETCH", f"< {clan} >开始查询...", externs.log_path)
        data_num = 0
        url_donation = url_0 + clans[clan]
        requests = urllib.request.Request(url = url_donation,headers = urls.HEADERS)
        response = urllib.request.urlopen(requests)
        content = response.read().decode("utf-8")
        tree = etree.HTML(content)
        table_donation = tree.xpath(externs.XPaths["InfoPlayersTable"])[0]
        for tr in table_donation:
            tr = BeautifulSoup(etree.tostring(tr,encoding='utf-8').decode('utf-8'),'html.parser')
            data = [clan]
            player_name = formal.keep_before_first_newline(tr.find('a',class_='block member_link').get_text().strip().replace('\u200c','').replace('\u2006','') )
            if filter and player_name not in players:
                continue
            data.append(player_name)
            donation = (int)(tr.find('td',class_='donations right aligned mobile-hide').get_text().strip().replace(',',''))
            data.append(donation)
            data_num = data_num + 1
            ws.append(data)
        now_row = now_row + data_num
        wb.save(filename = externs.outputFileLocation)
        formal.sort_xlsx_data(externs.outputFileLocation,sheet_name,start_row = start_row,end_row = now_row,sort_column = 3)
        wb = op.load_workbook(externs.outputFileLocation)
        ws = wb[sheet_name]
        ws.merge_cells(start_row = start_row,end_row = now_row,start_column = 1,end_column = 1)
        start_row = now_row + 1
        log.log("INFO", "FETCH", f"< {clan} >查询已完成！", externs.log_path)
    for row in ws.iter_rows(min_row = 2,max_row = ws.max_row,min_col=3,max_col=3):
        if row[0].value == 0:
            fill = PatternFill(fill_type = 'solid',fgColor = 'FF0000')
            ws.cell(row = row[0].row,column = 3).fill = fill
    wb.save(filename = externs.outputFileLocation)    
    log.log("INFO", "FETCH", "部落捐赠查询完成！", )
    log.log("INFO", "FETCH", "部落捐赠查询完成！", externs.log_path)
    return flag

def queryActivity(filter):
    flag = True
    sheet_name = formal.activitySheetName
    url_0 = urls.url_clan
    formal.creatSheet(operations.creat_activity_sheet())
    log.log("INFO", "FETCH", "输出创建完成,准备启动程序...", externs.log_path)
    wb = op.load_workbook(externs.outputFileLocation)
    ws = wb[sheet_name]
    data_num = 0
    now_row = 1
    start_row = 2
    log.log("INFO", "FETCH", "程序启动,正在运行...", externs.log_path)
    for clan in clans:
        log.log("INFO", "FETCH", f"< {clan} >开始查询...", externs.log_path)
        data_num = 0
        url_activity = url_0 + clans[clan]
        requests = urllib.request.Request(url = url_activity,headers = urls.HEADERS)
        response = urllib.request.urlopen(requests)
        content = response.read().decode("utf-8")
        tree = etree.HTML(content)
        table_activity = tree.xpath(externs.XPaths["InfoPlayersTable"])[0]
        for tr in table_activity:
            tr = BeautifulSoup(etree.tostring(tr,encoding='utf-8').decode('utf-8'),'html.parser')
            data = [clan]
            player_name = formal.keep_before_first_newline(tr.find('a',class_='block member_link').get_text().strip()).replace('\u200c','').replace('\u2006','')
            if filter and player_name not in players:
                continue
            data.append(player_name)
            activity = formal.convert_time_number(tr.find('div',class_='i18n_duration_short').get_text().strip())
            data.append(activity)
            data_num = data_num + 1
            ws.append(data)
        now_row = now_row + data_num
        wb.save(filename = externs.outputFileLocation)
        wb = op.load_workbook(externs.outputFileLocation)
        ws = wb[sheet_name]
        ws.merge_cells(start_row = start_row,end_row = now_row,start_column = 1,end_column = 1)
        start_row = now_row + 1
        log.log("INFO", "FETCH", f"< {clan} >查询已完成！", externs.log_path)
    for row in ws.iter_rows(min_row = 2,max_row = ws.max_row,min_col=3,max_col=3):
        if row[0].value >= externs.inactivity:
            fill = PatternFill(fill_type = 'solid',fgColor = 'FF0000')
            ws.cell(row = row[0].row,column = 3).fill = fill
        ws.cell(row = row[0].row,column = 3).value = formal.convert_time_from_number(row[0].value)
    wb.save(filename = externs.outputFileLocation)
    log.log("INFO", "FETCH", "部落活跃查询完成！", )
    log.log("INFO", "FETCH", "部落活跃查询完成！", externs.log_path)
    return flag

def queryLastMonthWar(filter):
    flag = True
    sheet_name = formal.lastMonthWarSheetName
    url_0 = urls.url_clan
    formal.creatSheet(operations.creat_last_month_war_sheet())
    log.log("INFO", "FETCH", "输出创建完成,准备启动程序...", externs.log_path)
    wb = op.load_workbook(externs.outputFileLocation)
    ws = wb[sheet_name]
    data_num = 0
    now_row = 1
    start_row = 2
    log.log("INFO", "FETCH", "程序启动,正在运行...", externs.log_path)
    for clan in clans:
        log.log("INFO", "FETCH", "开始配置驱动...", externs.log_path)
        edge_options = Options()
        edge_options.add_argument("--headless") 
        edge_options.add_argument("--disable-gpu") 
        edge_options.add_argument("--no-sandbox")  
        edge_options.add_argument("--disable-extensions")  
        edge_options.add_argument("--remote-debugging-port=0") 
        edge_options.add_argument("--ignore-certificate-errors")
        edge_options.add_argument("--ignore-ssl-errors")
        edge_options.add_experimental_option('excludeSwitches', ['enable-logging'])
        log.log("INFO", "FETCH", "开始配置日志输出...", externs.log_path)
        log_path = externs.Contributionslog_path + "\\" + clans[clan] + ".log"
        try:
            service = EdgeService(EdgeChromiumDriverManager().install(), service_args=["--verbose", " --log-path="+log_path])
        except Exception as e:
            log.log("CRITICAL", "FETCH", f"驱动初始化失败: {str(e)}", externs.log_path)
            log.log("CRITICAL", "FETCH", f"驱动初始化失败: {str(e)}", )
            flag = False
            return flag
        log.log("INFO", "FETCH",f"驱动启动项初始化完毕,日志输出为'{log_path}',准备启动...", externs.log_path)
        log.log("INFO", "FETCH", "驱动配置完成,准备启动...", externs.log_path)
        driver = webdriver.Edge(service = service,options = edge_options)
        log.log("INFO", "FETCH", "驱动已启动,准备开始查询...", externs.log_path)
        log.log("INFO", "FETCH", f"< {clan} >开始查询...", externs.log_path)
        data_num = 0
        url_last_month_war = url_0 + clans[clan] +"/war/analytics"
        driver.get(url_last_month_war)
        log.log("INFO", "FETCH", "驱动正在运行......在驱动正确退出前请不要结束进程,否则将导致数据不完整和其他可能的错误！", )
        log.log("INFO", "FETCH", "页面加载中...", externs.log_path)
        try:
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH,externs.XPaths["LastMonthWarTable1"]))
            )
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH,externs.XPaths["LastMonthWarTable2"]))
            )
            log.log("INFO", "FETCH", "页面加载完成,准备解析...", externs.log_path)
            log.log("INFO", "FETCH", "页面加载完成,准备解析...", )
            content = driver.page_source
        except:
            log.log("ERROR", "FETCH", "页面加载超时,驱动即将退出...", externs.log_path)
            log.log("ERROR", "FETCH", "页面加载超时,驱动即将退出...", )
            driver.quit()
            log.log("INFO", "FETCH", f"驱动已正确退出,具体信息请查看日志：'{log_path}'", externs.log_path)
            log.log("INFO", "FETCH", f"驱动已正确退出,具体信息请查看日志：'{log_path}'", )
            flag = False
            continue
        tree = etree.HTML(content)
        contributionsTable = tree.xpath(externs.XPaths["LastMonthWarTable1"])[0]
        length = len(contributionsTable)+1
        for i in range(1,length):
            player_name =  tree.xpath(externs.XPaths["LastMonthWarTable1"] + "/tr[" + str(i) + "]/td[1]/a")[0].text.strip().replace('\u200c','').replace('\u2006','')
            data = [clan, player_name]
            contributionList = []
            decksList = []
            all_contributions=0
            all_decks=0
            for j in range(5,9):
                contribution = tree.xpath(externs.XPaths["LastMonthWarTable1"] + "/tr[" + str(i) + "]/td[" + str(j) + "]")[0].text.strip().replace(" ","")
                if contribution == "":
                    contribution = 0
                contribution = (int)(contribution)
                all_contributions=all_contributions+contribution
                contributionList.append(contribution)
                deck = tree.xpath(externs.XPaths["LastMonthWarTable2"] + "/tr[" + str(i) + "]/td[" + str(j) + "]")[0].text.strip().replace(" ","")
                if deck == "":
                    deck = 0
                deck = (int)(deck)
                all_decks=all_decks+deck
                decksList.append(deck)
            contributionList.reverse()
            contributionList.append(all_contributions)
            decksList.reverse()
            decksList.append(all_decks)
            for deck,contribution in zip(decksList,contributionList):
                data.append(deck)
                data.append(contribution)
            participation = tree.xpath(externs.XPaths["LastMonthWarTable1"] + "/tr[" + str(i) + "]/td[3]")[0].text.strip().replace(" ","")
            data.append(participation)
            data_num = data_num + 1
            ws.append(data)
        now_row = now_row + data_num
        wb.save(filename = externs.outputFileLocation)
        formal.sort_xlsx_data(externs.outputFileLocation,sheet_name,start_row = start_row,end_row = now_row,sort_column = 12)
        wb = op.load_workbook(externs.outputFileLocation)
        ws = wb[sheet_name]
        ws.merge_cells(start_row = start_row,end_row = now_row,start_column = 1,end_column = 1)
        start_row = now_row + 1
        log.log("INFO", "FETCH", f"< {clan} >查询已完成！驱动准备退出...", externs.log_path)
        driver.quit()
        log.log("INFO", "FETCH", f"驱动已正确退出,具体信息请查看日志：'{log_path}'", externs.log_path)
    for row in ws.iter_rows(min_row = 2,max_row = ws.max_row,min_col=4,max_col=12):
        if row[0].value == 0:
            fill = PatternFill(fill_type = 'solid',fgColor = 'FF0000')
            ws.cell(row = row[0].row,column = 4).fill = fill
        if row[2].value == 0:
            fill = PatternFill(fill_type = 'solid',fgColor = 'FF0000')
            ws.cell(row = row[0].row,column = 6).fill = fill
        if row[4].value == 0:
            fill = PatternFill(fill_type = 'solid',fgColor = 'FF0000')
            ws.cell(row = row[0].row,column = 8).fill = fill
        if row[6].value == 0:
            fill = PatternFill(fill_type = 'solid',fgColor = 'FF0000')
            ws.cell(row = row[0].row,column = 10).fill = fill
        if row[8].value == 0:
            fill = PatternFill(fill_type = 'solid',fgColor = 'FF0000')
            ws.cell(row = row[0].row,column = 12).fill = fill
    wb.save(filename = externs.outputFileLocation)
    log.log("INFO", "FETCH", "上月部落战贡献查询完成！", )
    log.log("INFO", "FETCH", "上月部落战贡献查询完成！", externs.log_path)
    return flag

def queryLastMonthDonation(filter):
    flag = True
    sheet_name = formal.lastMonthDonationSheetName
    url_0 = urls.url_clan
    formal.creatSheet(operations.creat_last_month_donation_sheet())
    log.log("INFO", "FETCH", "输出创建完成,准备启动程序...", externs.log_path)
    wb = op.load_workbook(externs.outputFileLocation)
    ws = wb[sheet_name]
    data_num = 0
    now_row = 1
    start_row = 2
    log.log("INFO", "FETCH", "程序启动,正在运行...", externs.log_path)
    for clan in clans:
        log.log("INFO", "FETCH", "开始配置驱动...", externs.log_path)
        edge_options = Options()
        edge_options.add_argument("--headless") 
        edge_options.add_argument("--disable-gpu") 
        edge_options.add_argument("--no-sandbox")  
        edge_options.add_argument("--disable-extensions")  
        edge_options.add_argument("--remote-debugging-port=0") 
        edge_options.add_argument("--ignore-certificate-errors")
        edge_options.add_argument("--ignore-ssl-errors")
        edge_options.add_experimental_option('excludeSwitches', ['enable-logging'])
        log.log("INFO", "FETCH", "开始配置日志输出...", externs.log_path)
        log_path = externs.Donationslog_path + "\\" + clans[clan] + ".log"
        try:
            service = EdgeService(EdgeChromiumDriverManager().install(), service_args=["--verbose", " --log-path="+log_path])
        except Exception as e:
            log.log("CRITICAL", "FETCH", f"驱动初始化失败: {str(e)}", externs.log_path)
            log.log("CRITICAL", "FETCH", f"驱动初始化失败: {str(e)}", )
            flag = False
            return flag
        log.log("INFO", "FETCH", f"驱动启动项初始化完毕,日志输出为'{log_path}',准备启动...", externs.log_path)
        log.log("INFO", "FETCH", "驱动配置完成,准备启动...", externs.log_path)
        driver = webdriver.Edge(service = service,options = edge_options)
        log.log("INFO", "FETCH", "驱动已启动,准备开始查询...", externs.log_path)
        log.log("INFO", "FETCH", f"< {clan} >开始查询...", externs.log_path)
        data_num = 0
        url_last_month_donation = url_0 + clans[clan]+"/history"
        driver.get(url_last_month_donation)
        log.log("INFO", "FETCH", "驱动正在运行......在驱动正确退出前请不要结束进程,否则将导致数据不完整和其他可能的错误！", )
        log.log("INFO", "FETCH", "页面加载中...", externs.log_path)
        try:
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH,externs.XPaths["LastMonthDonation"]))
            )
            log.log("INFO", "FETCH", "页面加载完成,准备解析...", externs.log_path)
            log.log("INFO", "FETCH", "页面加载完成,准备解析...", )
            content = driver.page_source
        except:
            log.log("ERROR", "FETCH", "页面加载超时,驱动即将退出...", externs.log_path)
            log.log("ERROR", "FETCH", "页面加载超时,驱动即将退出...", )
            driver.quit()
            log.log("INFO", "FETCH", f"驱动已正确退出,具体信息请查看日志：'{log_path}'", externs.log_path)
            log.log("INFO", "FETCH", f"驱动已正确退出,具体信息请查看日志：'{log_path}'", )
            flag = False
            continue
        tree = etree.HTML(content)
        table_donation = tree.xpath(externs.XPaths["LastMonthDonation"])[0]
        length = len(table_donation)+1
        for i in range(1,length):
            player_name = tree.xpath(externs.XPaths["LastMonthDonation"] + "/tr[" + str(i) + "]/td[1]/a")[0].text.strip().replace('\u200c','').replace('\u2006','')
            data = [clan, player_name]
            donationList = []
            all_donations = 0
            for j in range(3,7):
                donation = tree.xpath(externs.XPaths["LastMonthDonation"] + "/tr[" + str(i) + "]/td[" + str(j) + "]/div")[0].text
                if donation is None:
                    donation = 0
                donation = (int)(donation)
                all_donations = all_donations + donation
                donationList.append(donation)
            donationList.reverse()
            donationList.append(all_donations)
            for donation in donationList:
                data.append(donation)
            data_num = data_num + 1
            ws.append(data)
        now_row = now_row + data_num
        wb.save(filename = externs.outputFileLocation)
        formal.sort_xlsx_data(externs.outputFileLocation,sheet_name,start_row = start_row,end_row = now_row,sort_column = 7)
        wb = op.load_workbook(externs.outputFileLocation)
        ws = wb[sheet_name]
        ws.merge_cells(start_row = start_row,end_row = now_row,start_column = 1,end_column = 1)
        start_row = now_row + 1
        log.log("INFO", "FETCH", f"< {clan} >查询已完成！驱动准备退出...", externs.log_path)
        driver.quit()
        log.log("INFO", "FETCH", f"驱动已正确退出,具体信息请查看日志：'{log_path}'", externs.log_path)
    for row in ws.iter_rows(min_row = 2,max_row = ws.max_row,min_col=3,max_col=7):
        if row[0].value == 0:
            fill = PatternFill(fill_type = 'solid',fgColor = 'FF0000')
            ws.cell(row = row[0].row,column = 3).fill = fill
        if row[1].value == 0:
            fill = PatternFill(fill_type = 'solid',fgColor = 'FF0000')
            ws.cell(row = row[0].row,column = 4).fill = fill
        if row[2].value == 0:
            fill = PatternFill(fill_type = 'solid',fgColor = 'FF0000')
            ws.cell(row = row[0].row,column = 5).fill = fill
        if row[3].value == 0:
            fill = PatternFill(fill_type = 'solid',fgColor = 'FF0000')
            ws.cell(row = row[0].row,column = 6).fill = fill
        if row[4].value == 0:
            fill = PatternFill(fill_type = 'solid',fgColor = 'FF0000')
            ws.cell(row = row[0].row,column = 7).fill = fill
    wb.save(filename = externs.outputFileLocation)
    log.log("INFO", "FETCH", "上月部落捐赠查询完成！", )
    log.log("INFO", "FETCH", "上月部落捐赠查询完成！", externs.log_path)
    return flag

def queryAndSort(filter):
    flag = True
    log.log("INFO", "FETCH", "准备查询前置信息...", externs.log_path)
    log.log("INFO", "FETCH", "准备查询前置信息...", )
    flag = queryLastMonthWar(filter) and queryLastMonthDonation(filter)
    if flag == False:
        log.log("ERROR", "FETCH", "前置查询失败,程序准备退出...", externs.log_path)
        log.log("ERROR", "FETCH", "前置查询失败,程序准备退出...", )
        return flag
    log.log("INFO", "FETCH", "前置查询完成,准备设定权重并启动程序...", externs.log_path)
    contribution_weight = externs.weightContribution
    donation_weight = externs.weightDonation
    log.log("INFO", "FETCH", "权重设定完毕,准备创建输出...", externs.log_path)
    formal.creatSheet(operations.creat_sort_sheet())
    log.log("INFO", "FETCH", "输出创建完成,准备启动程序...", externs.log_path)
    wb = op.load_workbook(externs.outputFileLocation)
    ws_1 = wb[formal.lastMonthWarSheetName]
    ws_2 = wb[formal.lastMonthDonationSheetName]
    ws_3 = wb[formal.sortedSheetName]
    results = []
    last_clan = ""
    for row_1 in ws_1.iter_rows(min_row=2, max_row=ws_1.max_row, min_col=2, max_col=2):
        results = [(ws_1.cell(row = row_1[0].row, column = 1).value)]
        if results[0]:
            last_clan = results[0]
        else:
            results[0] = last_clan
        str_value = row_1[0].value 
        results.append(str_value)
        for row_2 in ws_2.iter_rows(min_row=2, max_row=ws_2.max_row, min_col=2, max_col=2):
            if row_2[0].value == str_value:
                value_1 = ws_1.cell(row=row_1[0].row, column=12).value
                value_2 = ws_2.cell(row=row_2[0].row, column=7).value
                result = (value_1 * contribution_weight) + (value_2 * donation_weight)
                result = (int)(result)
                results.append(value_1)
                results.append(value_2)
                results.append(result)
                ws_3.append(results)
                break
    end = ws_3.max_row
    wb.save(filename = externs.outputFileLocation)
    log.log("INFO", "FETCH", "数据已保存,准备排序...", externs.log_path)
    formal.sort_xlsx_data(externs.outputFileLocation,formal.sortedSheetName,start_row = 2,end_row = end,sort_column = 5)
    wb = op.load_workbook(externs.outputFileLocation)
    ws = wb[formal.sortedSheetName]
    for row in ws.iter_rows(min_row = 2,max_row = ws.max_row,min_col = 3,max_col = 5):
        if row[0].value == 0:
            fill = PatternFill(fill_type = 'solid',fgColor = 'FF0000')
            ws.cell(row = row[0].row,column = 3).fill = fill
        if row[1].value == 0:
            fill = PatternFill(fill_type = 'solid',fgColor = 'FF0000')
            ws.cell(row = row[0].row,column = 4).fill = fill
        if row[2].value == 0:
            fill = PatternFill(fill_type = 'solid',fgColor = 'FF0000')
            ws.cell(row = row[0].row,column = 5).fill = fill
    log.log("INFO", "FETCH", "排序完成,准备保存权重信息...", externs.log_path)
    ws.append(['贡献权重','捐赠权重'])
    ws.append([contribution_weight,donation_weight])
    wb.save(filename = externs.outputFileLocation)
    log.log("INFO", "FETCH", "权重信息已保存,程序已退出", externs.log_path)
    log.log("INFO", "FETCH", "贡献计算与排序完成！", )
    log.log("INFO", "FETCH", "贡献计算与排序完成！", externs.log_path)
    return flag

def queryRecentChange():
    flag = True
    sheet_name = formal.recentChangeSheetName
    url_0 = urls.url_clan
    formal.creatSheet(operations.creat_recent_change_sheet())
    log.log("INFO", "FETCH", "输出创建完成,准备启动程序...", externs.log_path)
    wb = op.load_workbook(externs.outputFileLocation)
    ws = wb[sheet_name]
    data_num = 0
    now_row = 1
    start_row = 2
    log.log("INFO", "FETCH", "程序启动,正在运行...", externs.log_path)
    for clan in clans:
        log.log("INFO", "FETCH", f"< {clan} >开始查询...", externs.log_path)
        data_num = 0
        url_change = url_0 + clans[clan] + "/history/join-leave"
        requests = urllib.request.Request(url = url_change,headers = urls.HEADERS)
        response = urllib.request.urlopen(requests)
        content = response.read().decode("utf-8")
        tree = etree.HTML(content)
        players = tree.xpath(externs.XPaths["JoinLeaveTable"])[0]
        length = len(players)+1
        dict_change = {}
        for i in range(1,length):
            player = etree.tostring(players[i-1],encoding='utf-8').decode('utf-8')
            player_name = tree.xpath(externs.XPaths["JoinLeaveTable"] + "/a[" + str(i)+ "]/div/div/div[1]")[0].text.strip().replace('\u200c','').replace('\u2006','')
            if player_name not in dict_change:
                dict_change[player_name] = 0
            if "green plus icon" in player:
                dict_change[player_name] = dict_change[player_name] + 1
            elif "red minus icon" in player:
                dict_change[player_name] = dict_change[player_name] - 1
        in_num = 0
        out_num = 0
        for player in dict_change:
            data = [clan]
            if dict_change[player] == 0:
                continue
            elif dict_change[player] == 1:
                data.append(player)
                data.append("是")
                data.append("")
                data_num = data_num + 1
                in_num = in_num + 1
                ws.append(data)
            elif dict_change[player] == -1:
                data.append(player)
                data.append("")
                data.append("是")
                data_num = data_num + 1
                out_num = out_num + 1
                ws.append(data)
        ws.append([clan,"合计",in_num,out_num])
        url_all_members = url_0 + clans[clan]
        requests = urllib.request.Request(url = url_all_members,headers = urls.HEADERS)
        response = urllib.request.urlopen(requests)
        content = response.read().decode("utf-8")
        tree = etree.HTML(content)
        members = (int)(tree.xpath(externs.XPaths["InfoNumber"])[0].text.split('/')[0].strip())
        ws.append([clan,"目前成员数",members,members])
        data_num = data_num + 2
        now_row = now_row + data_num
        wb.save(filename = externs.outputFileLocation)
        wb = op.load_workbook(externs.outputFileLocation)
        ws = wb[sheet_name]
        ws.merge_cells(start_row = start_row,end_row = now_row,start_column = 1,end_column = 1)
        ws.merge_cells(start_row = now_row,start_column = 3,end_row = now_row,end_column = 4)
        start_row = now_row + 1
        log.log("INFO", "FETCH", f"< {clan} >查询已完成！", externs.log_path)
    for row in ws.iter_rows(min_row = 2,max_row = ws.max_row,min_col = 3,max_col = 4):
        if row[0].value == "是":
            ws.cell(row = row[0].row,column = 3).value = ""
            fill = PatternFill(fill_type = 'solid',fgColor = '538DD5')
            ws.cell(row = row[0].row,column = 3).fill = fill
        elif row[1].value == "是":
            ws.cell(row = row[1].row,column = 4).value = ""
            fill = PatternFill(fill_type = 'solid',fgColor = 'DA9694')
            ws.cell(row = row[0].row,column = 4).fill = fill
    wb.save(filename = externs.outputFileLocation)
    log.log("INFO", "FETCH", "部落成员变动查询完成！", )
    log.log("INFO", "FETCH", "部落成员变动查询完成！", externs.log_path)
    return flag