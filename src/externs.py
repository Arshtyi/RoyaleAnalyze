"""
该模块是 RoyaleAnalyze 项目的一部分,主要用于处理部落和玩家信息的读取和操作。
模块包含以下内容:
- 常量定义:包括工作表名称、文件路径、日志路径、权重、活动时间等。
- 函数定义:
    - getClansInformation(): 从指定路径的 Excel 文件中读取 'clansInformation' sheet 的内容,并返回一个字典。
    - getGroupPlayersInformation(): 从指定路径的 Excel 文件中读取 'groupPlayerInformation' sheet 的内容,并返回一个字典。
- 全局变量:
    - clans: 存储部落信息的字典,由 getClansInformation() 函数生成。
    - players: 存储玩家信息的字典,由 getGroupPlayersInformation() 函数生成。
    - operations: 操作码与操作描述的映射字典。
    - XPaths: 存储用于查询网页元素的 XPath 字典。
该模块的主要功能是从 Excel 文件中读取部落和玩家信息,并根据特定的规则进行处理和存储。

"""
import src.path as path
import openpyxl as op
import os
clansInformationSheetName = "Clans"
groupPlayerInformationSheetName = "Group"
contributionsSheetName = "Contribution"
donationsSheetName = "Donation"
activitySheetName = "Activity"
lastMonthWarSheetName = "LastMonthWar"
lastMonthDonationSheetName = "LastMonthDonation"
sortedSheetName = "Sorted"
recentChangeSheetName = "RecentChange"
inputClansInformationLocation = path.pathConcatenationForClansInformationTable()
inputGroupPlayerInformationLocation = path.pathConcatenationForGroupPlayerInformationTable()
outputFileLocation = path.pathConcatenationForOutputTable()
log_dir_path = path.log_dir_path()
log_path= path.log_path()
readme_path = path.readme_path()
Contributionslog_path = path.Contributionslog_path()
Donationslog_path = path.Donationslog_path()
FaultsLog_path = path.FaultsLog_path()
weightContribution = 0.25
weightDonation = 0.75
inactivity = 5 * 24 * 60
def getClansInformation():
    workbook = op.load_workbook(filename = inputClansInformationLocation)
    # 确保 sheet 名为 'clansInformation'
    if clansInformationSheetName not in workbook.sheetnames:
        raise ValueError(f"Excel 文件中没有名为 < {clansInformationSheetName} > 的 sheet")
    sheet = workbook[clansInformationSheetName]
    result = {}
    # 遍历从第二行开始的内容
    for row in sheet.iter_rows(min_row=2, max_col=2, values_only=True):
        key, value = row
        if key is not None:  # 确保 key 不为空
            result[key] = value
    workbook.close()
    return result
def getGroupPlayersInformation():
    if not os.path.exists(inputGroupPlayerInformationLocation):
        return {}
    wb = op.load_workbook(filename = inputGroupPlayerInformationLocation)
    if groupPlayerInformationSheetName not in wb.sheetnames:
        raise ValueError(f"Excel 文件中没有名为 < {groupPlayerInformationSheetName} > 的 sheet")
    ws = wb[groupPlayerInformationSheetName]
    result = {}
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=5, values_only=True):
        if row[0] in clans and (row[4] != 'y' and row[4] != 'Y' and row[4] != '是'):
            result[row[1]] = [row[0], row[2], row[3]]
    wb.close()
    return result
clans = getClansInformation()
players = getGroupPlayersInformation()
operations = {
    -1: "比对更新部落信息和玩家信息",
    0: "清除输出文件",
    1: "查询当前部落战贡献",
    2: "查询当前部落成员捐赠",
    3: "查询部落成员最近活跃情况",
    4: "查询近一月部落战贡献",
    5: "查询近一月捐赠",
    6: "根据4、5进行排序",
    7: "查询近一月成员变动",
    99: "格式化并退出"
}
XPaths = {
    "ClanName":"/html/body/div[3]/div[4]/div[2]/div[3]/div[1]/div[2]/div[1]/h1",
    "PlayerName":"/html/body/div[3]/div[4]/div[2]/div[3]/div[1]/div[2]/div[1]/h1", 
    "PlayerClan":"/html/body/div[3]/div[4]/div[2]/div[3]/div[1]/div[3]/div[4]/div[1]/div[2]/a",
    "WarTimeline":"/html/body/div[3]/div[4]/div[2]/div[4]/div[1]/div[1]/div[1]/ul",
    "WarPlayersTable":"/html/body/div[3]/div[4]/div[2]/div[4]/div[3]/table/tbody",
    "InfoPlayersTable":"/html/body/div[3]/div[4]/div[2]/div[4]/div[6]/table/tbody",
    "LastMonthWarTable1": "/html/body/div[3]/div[4]/div[2]/div[4]/div[2]/div/div/div/div[2]/div/table/tbody",
    "LastMonthWarTable2": "/html/body/div[3]/div[4]/div[2]/div[4]/div[4]/div/div/div/div[2]/div/table/tbody",
    "LastMonthDonation": "/html/body/div[3]/div[4]/div[2]/div[4]/div[3]/div/div[2]/div/table/tbody",
    "JoinLeaveTable":"/html/body/div[3]/div[4]/div[2]/div[4]",
    "InfoNumber":"/html/body/div[3]/div[4]/div[2]/div[4]/div[1]/div/div[2]/div[4]/div/div"
}