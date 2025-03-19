"""
formal.py
该模块是 RoyaleAnalyze 项目的一部分,主要负责处理 Excel 文件的创建、修改和格式化操作。模块中的函数用于清除工作表内容、检查文件存在性、删除工作表、创建特定格式的工作表、调整列宽、设置单元格对齐方式、排序数据以及时间格式转换等操作。
模块包含以下主要功能:
1. clearSheet(fileName, sheetName): 清除指定工作表的所有内容、格式和合并单元格,完全重置工作表。
2. checkFile(fileName): 检查当前目录下是否存在指定的 Excel 文件。
3. delSheetFromWorkbook(fileName, sheetName): 删除特定的工作表,如果工作表数量大于1则删除,否则清除内容。
4. keep_before_first_newline(input_string): 保留字符串中第一个换行符之前的内容。
5. creatSheet(operation): 根据操作类型创建特定格式的工作表,并添加表头。
6. calculateAdjustedWidth(value): 根据单元格内容计算宽度,处理中文繁体和宽字符的显示宽度。
7. adjustColumnWidth(sheet): 动态调整工作表的列宽,确保中文繁体和宽字符正确显示。
8. processExcel(fileName): 处理 Excel 文件,调整列宽和设置单元格居中。
9. sort_xlsx_data(file_path, sheet_name, start_row, end_row, sort_column): 按照给定的依据列对指定范围的行进行降序排序。
10. convert_time_number(input_str): 将时间字符串转换为总分钟数。
11. convert_time_from_number(total_minutes): 将总分钟数转换为时间字符串格式。
12. convert_time_format(input_str): 将时间字符串转换为带有中文单位的格式。
13. modify_line_in_file(file_path, line_number, new_content): 修改文件中特定行的内容。
该模块依赖于 openpyxl 库进行 Excel 文件操作,并使用 os 库进行文件系统操作。
"""
import src.externs as externs
import openpyxl as op
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import src.operations as operations
import os
import src.gettime as gettime
import src.externs as externs
import src.log as log
outputName = externs.outputFileLocation
contributionSheetName = externs.contributionsSheetName
donationSheetName = externs.donationsSheetName
activitySheetName = externs.activitySheetName
lastMonthWarSheetName = externs.lastMonthWarSheetName
lastMonthDonationSheetName = externs.lastMonthDonationSheetName
sortedSheetName = externs.sortedSheetName
recentChangeSheetName = externs.recentChangeSheetName
clansInformationSheetName = externs.clansInformationSheetName
def clearSheet(fileName, sheetName):
    """
    清除指定工作表的所有内容、格式和合并单元格,完全重置工作表.
    :param file_name: Excel 文件的名称
    :param sheet_name: 目标工作表的名称
    """
    wb = op.load_workbook(fileName)
    ws = wb[sheetName]
    # 取消所有合并单元格
    for merged_range in list(ws.merged_cells):
        ws.unmerge_cells(str(merged_range))  # 取消合并单元格
    # 清除所有单元格的内容和格式
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None  # 清空内容
            cell.style = 'Normal'  # 清除格式
    # 保存修改后的工作簿
    wb.save(fileName)

def checkFile(fileName):##当前目录下是否存在输出workbook
    if os.path.isfile(fileName):
        return True
    else:
        return False
    
def delSheetFromWorkbook(fileName,sheetName):##特定的sheet是否存在
    wb = op.load_workbook(filename = fileName)
    if sheetName in wb.sheetnames :##存在即删除
        if len(wb.sheetnames) > 1:##删或者清除
            del wb[sheetName]
        else :
            clearSheet(fileName,sheetName)
    wb.save(filename = fileName)
    ###不存在无事

def keep_before_first_newline(input_string):
    # 使用split方法按照 '\n' 切分字符串,取第一个部分
    return input_string.split('\n')[0]

def creatSheet(operation):
    if operations.creat_contribution_sheet() == operation:
        sheetName = contributionSheetName
        if checkFile(outputName):##输出文件是否存在
            ###存在则删除贡献sheet
            delSheetFromWorkbook(outputName,sheetName)
            wb = op.load_workbook(outputName)
            wb.create_sheet(title = sheetName)
        else :##不存在新建
            wb = op.Workbook()
            wb.create_sheet(title = sheetName)
        ws = wb[sheetName]
        ws.append(['部落','玩家','总使用卡组数','总袭击战船次数','总贡献',gettime.get_current_time()])##表头
        wb.save(filename = outputName)
    elif operations.creat_donation_sheet() == operation:
        sheetName = donationSheetName
        if checkFile(outputName): ##输出文件是否存在
            delSheetFromWorkbook(outputName,sheetName)
            wb = op.load_workbook(outputName)
            wb.create_sheet(title = sheetName)
        else: ##不存在新建
            wb = op.Workbook()
            wb.create_sheet(title = sheetName)
        ws = wb[sheetName]
        ws.append(['部落','玩家','本周捐赠',gettime.get_current_time()])##表头
        wb.save(filename = outputName)
    elif operations.creat_activity_sheet() == operation:
        sheetName = activitySheetName
        if checkFile(outputName):##输出文件是否存在
            delSheetFromWorkbook(outputName,sheetName)
            wb = op.load_workbook(outputName)
            wb.create_sheet(title = sheetName)
        else:##不存在新建
            wb = op.Workbook()
            wb.create_sheet(title = sheetName)
        ws = wb[sheetName]
        ws.append(['部落','玩家','当前最后活跃时间',gettime.get_current_time()])##表头
        wb.save(filename = outputName)
    elif operations.creat_last_month_war_sheet() == operation:
        sheetName = externs.lastMonthWarSheetName
        if checkFile(outputName):
            delSheetFromWorkbook(outputName,sheetName)
            wb = op.load_workbook(outputName)
            wb.create_sheet(title = sheetName)
        else:
            wb = op.Workbook()
            wb.create_sheet(title = sheetName)
        ws = wb[sheetName]
        ws.append(['部落','玩家','第一周使用卡组数','第一周贡献','第二周使用卡组数','第二周贡献','第三周使用卡组数','第三周贡献','第四周使用卡组数','第四周贡献','近一月使用卡组数','近一月贡献','近期参与的河道竞速',gettime.get_current_time()])
        wb.save(filename = outputName)
    elif operations.creat_last_month_donation_sheet() == operation:
        sheetName = externs.lastMonthDonationSheetName
        if checkFile(outputName):
            delSheetFromWorkbook(outputName,sheetName)
            wb = op.load_workbook(outputName)
            wb.create_sheet(title = sheetName)
        else:
            wb = op.Workbook()
            wb.create_sheet(title = sheetName)
        ws = wb[sheetName]
        ws.append(['部落','玩家','第一周捐赠','第二周捐赠','第三周捐赠','第四周捐赠','近一月捐赠',gettime.get_current_time()])
        wb.save(filename = outputName)
    elif operations.creat_sort_sheet() == operation:
        sheetName = externs.sortedSheetName
        if checkFile(outputName):
            delSheetFromWorkbook(outputName,sheetName)
            wb = op.load_workbook(outputName)
            wb.create_sheet(title = sheetName)
        else:
            wb = op.Workbook()
            wb.create_sheet(title = sheetName)
        ws = wb[sheetName]
        ws.append(['部落','玩家','上月贡献','上月捐赠','贡献',gettime.get_current_time()])
        wb.save(filename = outputName)
    elif operations.creat_recent_change_sheet() == operation:
        sheetName = "RecentChange"
        if checkFile(outputName):
            delSheetFromWorkbook(outputName,sheetName)
            wb = op.load_workbook(outputName)
            wb.create_sheet(title = sheetName)
        else:
            wb = op.Workbook()
            wb.create_sheet(title = sheetName)
        ws = wb[sheetName]
        ws.append(['部落','玩家','加入','退出',gettime.get_current_time()])
        wb.save(filename = outputName)
    else :
        log.log("ERROR", "FORMAL", "Undefined Query Type, Please Check Input Validity.", externs.log_path)

def calculateAdjustedWidth(value):
    """
    根据单元格内容计算宽度,处理中文繁体和宽字符的显示宽度.
    :param value: 单元格内容
    :return: 调整后的宽度(适配 Excel 的列宽标准)
    """
    if value is None:
        return 0
    value = str(value)
    width = 0
    for char in value:
        # 宽字符 (中文、繁体、日文、韩文、emoji等)
        if '\u4e00' <= char <= '\u9fff' or ord(char) > 127:
            width += 2  # 宽字符占用2单位宽度
        else:
            width += 1  # 普通字符占用1单位宽度
    return width

def adjustColumnWidth(sheet):
    """
    动态调整工作表的列宽,确保中文繁体和宽字符正确显示.
    :param sheet: 当前工作表
    """
    log.log("TRACE", "FORMAL", "开始调整列宽...", externs.log_path)
    for col in sheet.columns:
        max_width = 0
        col_letter = get_column_letter(col[0].column)  # 获取列号字母
        for cell in col:
            if cell.value is not None:
                adjusted_width = calculateAdjustedWidth(cell.value)
                max_width = max(max_width, adjusted_width)
        sheet.column_dimensions[col_letter].width = max_width * 1.2
    log.log("TRACE", "FORMAL", "列宽调整完成...", externs.log_path)

def processExcel(fileName):
    """
    处理Excel文件;
    1. 检查当前目录是否存在指定的Excel文件.
       - 如果不存在,抛出错误提示.
    2. 删除名为 "Sheet" 的工作表(如果 sheetnames 数量大于1).
    3. 对剩下的每个工作表;
        - 调整列宽.
        - 设置单元格内容居中.
    :param file_name: str, Excel 文件名.
    """
    # 检查当前目录是否存在指定文件
    if not os.path.exists(fileName):
        log.log("ERROR", "FORMAL", f"File '{fileName}' does not exist! Please check the file name and path.", externs.log_path)
        return
    wb = op.load_workbook(fileName)
    # 删除 "Sheet" 工作表
    if len(wb.sheetnames) > 1 and "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    # 遍历每个工作表
    for sheet in wb.worksheets:
        if sheet.title[-1] == "1":
            delSheetFromWorkbook(fileName,sheet.title)
        if len(wb.sheetnames) == 0:
            os.remove(fileName)
        # 调整列宽
        adjustColumnWidth(sheet)
        # 设置单元格居中
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
    # 保存更改
    wb.save(fileName)
    log.log("INFO", "FORMAL", f"格式化完成，文件输出为: '{fileName}'", externs.log_path)
    log.log("INFO", "FORMAL", f"格式化完成，文件输出为: '{fileName}'")

def sort_xlsx_data(file_path, sheet_name, start_row, end_row, sort_column):
    """
    按照给定的依据列对指定范围的行进行降序排序,
    其中,依据列内容为整数,且不能直接交换数据.

    :param file_path: Excel文件路径
    :param sheet_name: 工作表名
    :param start_row: 起始行
    :param end_row: 终止行
    :param sort_column: 排序依据列(从1开始,1表示第一列)
    :return: 排序后的数据(二维列表)
    """
    log.log("TRACE", "FORMAL", f"开始对工作表 '{sheet_name}' 的数据进行排序...", externs.log_path)
    wb = op.load_workbook(file_path)
    sheet = wb[sheet_name]
    rows_to_sort = []
    for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=sheet.max_column, values_only=False):
        row_data = [cell.value for cell in row]
        rows_to_sort.append(row_data)
    def sort_key(value):
        if value is None:
            return float('-inf')
        return value
    rows_to_sort.sort(key=lambda x: sort_key(x[sort_column - 1]), reverse=True)
    for i, row_data in enumerate(rows_to_sort, start=start_row):
        for j, value in enumerate(row_data):
            sheet.cell(row=i, column=j + 1, value=value)
    wb.save(file_path)
    log.log("TRACE", "FORMAL", f"工作表 '{sheet_name}' 的数据已排序完成.", externs.log_path)
    log.log("TRACE", "FORMAL", f"工作表 '{sheet_name}' 的数据已排序完成.")

def convert_time_number(input_str):
    time_values = {"w": 0, "d": 0, "h": 0, "m": 0}
    parts = input_str.split()
    for part in parts:
        if part.endswith("w"):
            time_values["w"] = int(part[:-1])
        elif part.endswith("d"):
            time_values["d"] = int(part[:-1])
        elif part.endswith("h"):
            time_values["h"] = int(part[:-1])
        elif part.endswith("m"):
            time_values["m"] = int(part[:-1])
    total_minutes = time_values["w"] * 7 * 24 * 60 + time_values["d"] * 24 * 60 + time_values["h"] * 60 + time_values["m"]
    log.log("INFO", "FORMAL", f"Converted time string '{input_str}' to total minutes: {total_minutes}", externs.log_path)
    return total_minutes

def convert_time_from_number(total_minutes):
    time_units = {"w": "周", "d": "天", "h": "小时", "m": "分钟"}
    time_values = {"w": 0, "d": 0, "h": 0, "m": 0}
    time_values["w"], total_minutes = divmod(total_minutes, 7 * 24 * 60)
    time_values["d"], total_minutes = divmod(total_minutes, 24 * 60)
    time_values["h"], time_values["m"] = divmod(total_minutes, 60)
    output = []
    if time_values["w"] > 0:
        output.append(f"{time_values['w']}周")
    if time_values["d"] > 0 or (time_values["w"] > 0 and (time_values["h"] > 0 or time_values["m"] > 0)):
        output.append(f"{time_values['d']}天")
    if time_values["h"] > 0 or ((time_values["w"] > 0 or time_values["d"] > 0) and time_values["m"] > 0):
        output.append(f"{time_values['h']}时")
    if time_values["m"] > 0:
        output.append(f"{time_values['m']}分")

    return "".join(output)

def convert_time_format(input_str):
    time_units = {"w": "周", "d": "天", "h": "小时", "m": "分钟"}
    time_values = {"w": 0, "d": 0, "h": 0, "m": 0}
    parts = input_str.split()
    for part in parts:
        if part.endswith("w"):
            time_values["w"] = int(part[:-1])
        elif part.endswith("d"):
            time_values["d"] = int(part[:-1])
        elif part.endswith("h"):
            time_values["h"] = int(part[:-1])
        elif part.endswith("m"):
            time_values["m"] = int(part[:-1])
    output = []
    if time_values["w"] > 0:
        output.append(f"{time_values['w']}周")
    if time_values["d"] > 0 or (time_values["w"] > 0 and (time_values["h"] > 0 or time_values["m"] > 0)):
        output.append(f"{time_values['d']}天")
    if time_values["h"] > 0 or ((time_values["w"] > 0 or time_values["d"] > 0) and time_values["m"] > 0):
        output.append(f"{time_values['h']}时")
    if time_values["m"] > 0:
        output.append(f"{time_values['m']}分")

    return "".join(output)

# # 下面的函数已经弃用
# def modify_line_in_file(file_path, line_number, new_content):
#     # 打开文件读取内容
#     with open(file_path, 'r') as file:
#         lines = file.readlines()
#     # 检查行号是否有效
#     if line_number <= 0 or line_number > len(lines):
#         return
#     # 修改特定的行
#     lines[line_number - 1] = new_content + '\n'  # 需要换行符来保证行格式
#     # 打开文件进行写入
#     with open(file_path, 'w') as file:
#         file.writelines(lines)